from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
from openpyxl.styles import Font
import shutil
import os
from datetime import datetime
import tempfile
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D

app = Flask(__name__)

template_file_id = "1v0kQrMN2N83NJVjGgBBsS7OZ2nZuyQRq"

special_fields = {
    "工程編號": (6, 2, 4, "37/2024/DVPS"),
    "工程名稱": (7, 2, 4, "黑沙馬路行人道優化工程(第二期)"),
    "文件編號": (6, 8, 8, "") 
}
regular_fields = [
    ("報批之材料", 11, 3),
    ("牌子(如有)", 12, 3),
    ("預算表之項目編號", 11, 7),
    ("型號", 12, 6),
    ("貨期", 13, 6),
    ("數量", 14, 6),
    # Remove "附件" from here
]

attachment_fields = [
    ("附件", 15, 6)  # Add "附件" here
]

material_type_checkboxes = [
    ("結構", 7, 6),
    ("供水", 8, 6),
    ("建築", 7, 8),
    ("電氣", 8, 8),
    ("排水", 7, 10),
    ("其他", 8, 10)
]
material_status_checkboxes = [
    ("與設計相同", 13, 1),
    ("與標書相同", 14, 1),
    ("與後加工程建議書相同", 15, 1),
    ("同等質量", 16, 1),
    ("替換材料", 17, 1),
    ("原設計沒有指定", 18, 1)
]
attachment_type_checkboxes = [
    ("樣板", 16, 5),
    ("目錄", 17, 5),
    ("來源證", 16, 7),
    ("其他", 17, 7)
]

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Process form data and generate Excel file
        temp_file = generate_excel(request.form)
        file_name = request.form.get('檔案名稱', '材料報批表_filled.xlsx')
        if not file_name.endswith('.xlsx'):
            file_name += '.xlsx'
        return send_file(temp_file, as_attachment=True, download_name=file_name)
    return render_template('form.html', 
                           special_fields=special_fields,
                           regular_fields=regular_fields,
                           material_type_checkboxes=material_type_checkboxes,
                           material_status_checkboxes=material_status_checkboxes,
                           attachment_type_checkboxes=attachment_type_checkboxes,
                           attachment_fields=attachment_fields)  # Add this line

def generate_excel(form_data):
    # Create a temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()

    # Copy the local template file
    shutil.copy('forms/材料報批表.xlsx', temp_file.name)

    # Load the copied file
    wb = load_workbook(temp_file.name)
    wb._external_links = []
    ws = wb.active

    # Add the logo
    img = Image('images/logo.jpg')
    
    # Set the size of the image (convert cm to EMUs)
    width_emu = int(4.09 * 360000)  # 4.09 cm
    height_emu = int(1.62 * 360000)  # 1.62 cm

    # Create an OneCellAnchor for the image
    marker = AnchorMarker(col=0, colOff=133350, row=1, rowOff=38100)  # 14 pixels ≈ 133350 EMU, 4 pixels ≈ 38100 EMU
    size = XDRPositiveSize2D(cx=width_emu, cy=height_emu)
    img.anchor = OneCellAnchor(_from=marker, ext=size)

    # Add the image to the worksheet
    ws.add_image(img)

    # Fill in the special fields
    for field, (row, start_col, end_col, _) in special_fields.items():
        merged_cell = ws.cell(row=row, column=start_col)
        merged_cell.value = form_data.get(field, '')
        if field == "工程名稱":
            merged_cell.font = Font(size=10)
    
    # Fill in the regular fields
    for field, row, col in regular_fields:
        ws.cell(row=row, column=col, value=form_data.get(field, ''))

    # Fill in the attachment fields
    for field, row, col in attachment_fields:
        ws.cell(row=row, column=col, value=form_data.get(field, ''))

    # Fill in the checkboxes
    for checkboxes in [material_type_checkboxes, material_status_checkboxes, attachment_type_checkboxes]:
        for field, row, col in checkboxes:
            ws.cell(row=row, column=col, value="☑" + field if field in form_data else "□" + field)

    # Fill in the date
    ws.cell(row=21, column=7, value=form_data.get('日期', datetime.now().strftime("%Y/%m/%d")))

    # Save the workbook
    wb.save(temp_file.name)

    return temp_file.name

if __name__ == "__main__":
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)