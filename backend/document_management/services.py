# backend/document_management/services.py

import os
import base64
import uuid
import logging
import subprocess
import platform
import tempfile
import shutil
from datetime import datetime
from typing import Optional, Dict, Any, List

import boto3
from botocore.exceptions import ClientError
from flask import current_app
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from backend.db import mongo

logger = logging.getLogger(__name__)


def get_s3_client():
    """
    建立並回傳 boto3 S3 client。
    """
    aws_access_key = os.environ.get("AWS_ACCESS_KEY_ID", "")
    aws_secret_key = os.environ.get("AWS_SECRET_ACCESS_KEY", "")
    bucket_region = os.environ.get("AWS_DEFAULT_REGION", "ap-southeast-1")

    if not aws_access_key or not aws_secret_key:
        logger.warning("AWS_ACCESS_KEY_ID 或 AWS_SECRET_ACCESS_KEY 未設置，S3 上傳可能失敗。")

    session = boto3.session.Session(
        aws_access_key_id=aws_access_key,
        aws_secret_access_key=aws_secret_key,
        region_name=bucket_region
    )
    return session.client("s3")


def _upload_file_to_s3(base64_file: str, original_filename: str) -> str:
    """
    將 base64 編碼的檔案內容上傳到 S3，回傳 S3 key。
    """
    bucket_name = os.environ.get("AWS_S3_BUCKET", "")
    if not bucket_name:
        raise ValueError("AWS_S3_BUCKET not set in environment.")

    s3_client = get_s3_client()

    ext = ""
    if "." in original_filename:
        ext = original_filename.split(".")[-1]
    unique_id = str(uuid.uuid4())
    s3_key = f"documents/{unique_id}.{ext}" if ext else f"documents/{unique_id}"

    file_bytes = base64.b64decode(base64_file)
    try:
        s3_client.put_object(
            Bucket=bucket_name,
            Key=s3_key,
            Body=file_bytes
        )
    except ClientError as e:
        logger.error("Failed to upload to S3: %s", e)
        raise

    return s3_key


# ------------------------------------------------------------
#  通用 CRUD (documents collection)
# ------------------------------------------------------------

def create_document(data: Dict[str, Any]) -> str:
    title = data.get("title") or "Untitled"
    doc_type = data.get("doc_type") or "UNKNOWN"
    project_id = data.get("project_id")
    description = data.get("description", "")

    if project_id is None:
        raise ValueError("Missing project_id")

    doc = {
        "title": title,
        "doc_type": doc_type,
        "project_id": project_id,
        "description": description,
        "versions": [],
        "daily_report_id": data.get("daily_report_id"),
        "daily_report_data": data.get("daily_report_data"),
        "created_at": datetime.now(),
        "updated_at": datetime.now()
    }
    result = mongo.db["documents"].insert_one(doc)
    return str(result.inserted_id)


def get_document(doc_id: str) -> Optional[dict]:
    doc = mongo.db["documents"].find_one({"_id": mongo.db["documents"].object_id(doc_id)})
    if not doc:
        return None
    doc["id"] = str(doc["_id"])
    del doc["_id"]
    return doc


def update_document(doc_id: str, data: Dict[str, Any]) -> bool:
    update_fields = {}
    if "title" in data:
        update_fields["title"] = data["title"]
    if "doc_type" in data:
        update_fields["doc_type"] = data["doc_type"]
    if "project_id" in data:
        update_fields["project_id"] = data["project_id"]
    if "description" in data:
        update_fields["description"] = data["description"]

    if not update_fields:
        return False

    update_fields["updated_at"] = datetime.now()

    result = mongo.db["documents"].update_one(
        {"_id": mongo.db["documents"].object_id(doc_id)},
        {"$set": update_fields}
    )
    return (result.modified_count > 0)


def delete_document(doc_id: str) -> bool:
    doc = mongo.db["documents"].find_one({"_id": mongo.db["documents"].object_id(doc_id)})
    if not doc:
        return False

    # 刪除所有版本檔案 (S3)
    s3_client = get_s3_client()
    bucket_name = os.environ.get("AWS_S3_BUCKET", "")
    for ver in doc.get("versions", []):
        s3_key = ver.get("s3_key", "")
        if s3_key:
            try:
                s3_client.delete_object(Bucket=bucket_name, Key=s3_key)
            except ClientError as e:
                logger.warning("Failed to delete S3 key=%s: %s", s3_key, e)

    mongo.db["documents"].delete_one({"_id": doc["_id"]})
    return True


def list_documents(project_id: str = "", doc_type: str = "") -> List[dict]:
    query = {}
    if project_id:
        try:
            query["project_id"] = int(project_id)
        except:
            pass
    if doc_type:
        query["doc_type"] = doc_type

    cursor = mongo.db["documents"].find(query).sort("created_at", -1)
    results = []
    for d in cursor:
        d["id"] = str(d["_id"])
        del d["_id"]
        results.append(d)
    return results


# ------------------------------------------------------------
#  版本管理
# ------------------------------------------------------------

def add_document_version(doc_id: str, data: Dict[str, Any]) -> str:
    doc = mongo.db["documents"].find_one({"_id": mongo.db["documents"].object_id(doc_id)})
    if not doc:
        raise ValueError("Document not found")

    version_note = data.get("version_note", "")
    filename = data.get("filename", "")
    base64_file = data.get("base64_file", "")

    new_version_id = str(uuid.uuid4())
    s3_key = ""

    if base64_file:
        s3_key = _upload_file_to_s3(base64_file, filename)

    new_version = {
        "version_id": new_version_id,
        "version_note": version_note,
        "filename": filename,
        "s3_key": s3_key,
        "created_at": datetime.now()
    }

    mongo.db["documents"].update_one(
        {"_id": doc["_id"]},
        {
            "$push": {"versions": new_version},
            "$set": {"updated_at": datetime.now()}
        }
    )

    return new_version_id


def get_document_versions(doc_id: str) -> List[dict]:
    doc = mongo.db["documents"].find_one({"_id": mongo.db["documents"].object_id(doc_id)})
    if not doc:
        raise ValueError("Document not found")
    return doc.get("versions", [])


# ------------------------------------------------------------
# 產生/填寫 XLSX & PDF (Daily Report)
# ------------------------------------------------------------

def generate_diary_xlsx_only(site_diary_doc: Dict[str, Any]) -> str:
    # 取得 project 資料
    project_id = site_diary_doc.get("project_id")
    project_doc = mongo.db.projects.find_one({"id": project_id})
    if not project_doc:
        raise ValueError(f"Project not found in MongoDB (id={project_id}).")

    dr_data = site_diary_doc.get("daily_report_data", {})
    report_date = dr_data.get("report_date", None)
    date_str = report_date if report_date else ""

    project_name = project_doc.get("name", "")
    project_job_number = project_doc.get("job_number", "")
    contractor_name = project_doc.get("contractor", "")

    day_count_str = str(dr_data.get("day_count", ""))
    summary_str = dr_data.get("summary", "")
    weather_morning = dr_data.get("weather_morning", "")
    weather_noon = dr_data.get("weather_noon", "")

    start_date = project_doc.get("start_date")
    start_date_str = start_date.strftime("%Y-%m-%d") if start_date else ""

    duration_days = project_doc.get("duration_days")
    duration_type = project_doc.get("duration_type")
    if duration_days:
        if duration_type == "business":
            duration_str = f"{duration_days}工作天"
        else:
            duration_str = f"{duration_days}天"
    else:
        duration_str = ""

    worker_map = {
        "地盤總管": "D19",
        "工程師":  "D20",
        "管工":    "D21",
        "平水員":  "D22",
        "燒焊焊工": "D23",
        "機手":    "D24",
        "泥水工":  "D25",
        "紮鐵工":  "D26",
        "木板工":  "D27",
        "電工":    "D28",
        "水喉工":  "D29",
        "雜工":    "D30",
    }
    machine_map = {
        "挖掘機": "G19",
        "發電機": "G20",
        "風機":   "G21",
        "泥頭車": "G22",
        "吊機":   "G23",
        "機炮":   "G24",
        "屈鐵機": "G25",
        "風車鋸": "G26",
    }

    workers_dict = dr_data.get("workers", {})
    machines_dict = dr_data.get("machines", {})

    template_path = os.path.join(
        current_app.root_path,
        'report_templates',
        'daily_report.xlsx'
    )
    if not os.path.isfile(template_path):
        raise FileNotFoundError(f"daily_report.xlsx not found: {template_path}")

    temp_dir = tempfile.mkdtemp(prefix="diary_xlsx_")
    date_str_for_filename = date_str.replace("-", "") if date_str else "noDate"
    filled_xlsx_path = os.path.join(
        temp_dir,
        f"{date_str_for_filename}_daily_report.xlsx"
    )

    shutil.copy(template_path, filled_xlsx_path)

    wb = load_workbook(filled_xlsx_path)
    # Sheet1
    if "每日施工進度報告表" in wb.sheetnames:
        sh1 = wb["每日施工進度報告表"]
        sh1["C2"].value = date_str
        sh1["C2"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["K2"].value = start_date_str
        sh1["K2"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["D4"].value = project_name
        sh1["D4"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["D5"].value = project_job_number
        sh1["D5"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["D6"].value = contractor_name
        sh1["D6"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["E7"].value = weather_morning
        sh1["E7"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["H7"].value = weather_noon
        sh1["H7"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["D8"].value = duration_str
        sh1["D8"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["J8"].value = day_count_str
        sh1["J8"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["B10"].value = summary_str
        sh1["B10"].alignment = Alignment(horizontal='center', vertical='center')

        for w_type, cell_ref in worker_map.items():
            qty = workers_dict.get(w_type, 0)
            sh1[cell_ref].value = str(qty) if qty else ""
            sh1[cell_ref].alignment = Alignment(horizontal='center', vertical='center')

        for m_type, cell_ref in machine_map.items():
            qty = machines_dict.get(m_type, 0)
            sh1[cell_ref].value = str(qty) if qty else ""
            sh1[cell_ref].alignment = Alignment(horizontal='center', vertical='center')

        sh1.print_area = "A1:L46"

    # Sheet2
    if "每日本地工人及外地勞工施工人員紀錄表" in wb.sheetnames:
        sh2 = wb["每日本地工人及外地勞工施工人員紀錄表"]
        b3_date = date_str.replace("-", "/") if date_str else ""
        sh2["B3"].value = b3_date
        sh2["B3"].alignment = Alignment(horizontal='center', vertical='center')

        sh2["B6"].value = project_name
        sh2["B6"].alignment = Alignment(horizontal='center', vertical='center')

        sh2["B7"].value = project_job_number
        sh2["B7"].alignment = Alignment(horizontal='center', vertical='center')

        sh2["B8"].value = contractor_name
        sh2["B8"].alignment = Alignment(horizontal='center', vertical='center')

        total_workers_sum = sum(workers_dict.values()) if workers_dict else 0
        total_machines_sum = sum(machines_dict.values()) if machines_dict else 0
        if (total_workers_sum + total_machines_sum) == 0:
            for row in range(11, 26):
                for col in range(2, 5):
                    sh2.cell(row=row, column=col, value="☐")
            for row in range(28, 33):
                for col in range(2, 4):
                    sh2.cell(row=row, column=col, value="☐")
            sh2["C33"].value = 0
            sh2["C34"].value = 0

        sh2.print_area = "A1:E40"

    wb.save(filled_xlsx_path)
    wb.close()

    return filled_xlsx_path


def generate_diary_pdf_sheet(site_diary_doc: Dict[str, Any], sheet_name: str) -> str:
    xlsx_path = generate_diary_xlsx_only(site_diary_doc)
    temp_dir = os.path.dirname(xlsx_path)

    dr_data = site_diary_doc.get("daily_report_data", {})
    raw_date = dr_data.get("report_date", "")
    date_str = raw_date.replace("-", "") if raw_date else "noDate"

    if sheet_name == 'sheet1':
        target_sheet_name = "每日施工進度報告表"
        pdf_filename = f"{date_str}每日施工進度報告表.pdf"
    else:
        target_sheet_name = "每日本地工人及外地勞工施工人員紀錄表"
        pdf_filename = f"{date_str}施工人員紀錄表.pdf"

    pdf_path = os.path.join(temp_dir, pdf_filename)

    single_sheet_xlsx = os.path.join(temp_dir, f"{date_str}_{sheet_name}_only.xlsx")
    shutil.copy(xlsx_path, single_sheet_xlsx)

    wb = load_workbook(single_sheet_xlsx)
    if target_sheet_name not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"指定工作表 '{target_sheet_name}' 不存在於模板。")

    for sn in wb.sheetnames:
        if sn != target_sheet_name:
            wb.remove(wb[sn])
    wb.save(single_sheet_xlsx)
    wb.close()

    lo_exec = get_libreoffice_cmd()
    input_filter = "Calc Office Open XML"

    # ★ 調整/嘗試更多字型嵌入設定：EmbedStandardFonts, UseTaggedPDF 等
    #   也可嘗試 EmbedAllFonts=true 或 ExportEmbedFonts=true，視實際 LibreOffice 版本支援情況而定
    pdf_filter_options = "SelectPdfVersion=1;EmbedStandardFonts=true;UseTaggedPDF=true"
    output_filter = f"pdf:calc_pdf_Export:{pdf_filter_options}"

    if platform.system().lower().startswith("win"):
        cmd = (
            f'"{lo_exec}" --headless --convert-to "{output_filter}" '
            f'--infilter="{input_filter}" "{single_sheet_xlsx}" --outdir "{temp_dir}"'
        )
        try:
            result = subprocess.run(cmd, shell=True, check=True, capture_output=True)
        except subprocess.CalledProcessError as e:
            err_msg = e.stderr.decode('utf-8', errors='replace') if e.stderr else str(e)
            raise RuntimeError(f"LibreOffice PDF conversion failed (Windows): {err_msg}") from e
    else:
        cmd_list = [
            lo_exec, "--headless",
            "--convert-to", output_filter,
            f'--infilter={input_filter}',
            single_sheet_xlsx,
            "--outdir", temp_dir
        ]
        try:
            result = subprocess.run(cmd_list, shell=False, check=True, capture_output=True)
        except subprocess.CalledProcessError as e:
            err_msg = e.stderr.decode('utf-8', errors='replace') if e.stderr else str(e)
            raise RuntimeError(f"LibreOffice PDF conversion failed: {err_msg}") from e

    generated_pdf_name = os.path.splitext(single_sheet_xlsx)[0] + ".pdf"
    if os.path.isfile(generated_pdf_name):
        os.rename(generated_pdf_name, pdf_path)
    else:
        raise RuntimeError(f"PDF 轉檔失敗: {generated_pdf_name} 不存在")

    if not os.path.isfile(pdf_path):
        raise RuntimeError(f"PDF 轉檔失敗: {pdf_path} 不存在")

    return pdf_path


def get_libreoffice_cmd() -> str:
    custom_path = os.environ.get("LIBREOFFICE_PATH")
    if custom_path:
        if os.path.isfile(custom_path):
            return custom_path
        soffice_path = os.path.join(custom_path, "soffice")
        libreoffice_path = os.path.join(custom_path, "libreoffice")
        if os.path.isfile(soffice_path):
            return soffice_path
        if os.path.isfile(libreoffice_path):
            return libreoffice_path
        raise RuntimeError(f"LIBREOFFICE_PATH 無效: {custom_path}")

    from shutil import which
    for candidate in ("soffice", "libreoffice"):
        exe_path = which(candidate)
        if exe_path:
            return exe_path

    raise RuntimeError(
        "無法自動找到 libreoffice/soffice！請安裝 LibreOffice 或設定 LIBREOFFICE_PATH"
    )
