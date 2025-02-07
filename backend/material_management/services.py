# backend/material_management/services.py
import os
import tempfile
import shutil
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from flask import current_app

from .constants import (
    special_fields,
    regular_fields,
    attachment_fields,
    material_type_checkboxes,
    material_status_checkboxes,
    attachment_type_checkboxes
)

def generate_excel(form_data):
    """
    依照使用者表單資料，填入並產生材料報批的 Excel 檔。
    回傳產生之暫存檔路徑。
    """
    # 複製模板到暫存檔
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()

    # 模板路徑
    template_path = os.path.join(
        current_app.root_path, 
        'material_management', 
        'templates', 
        'material_template.xlsx'
    )
    shutil.copy(template_path, temp_file.name)

    wb = load_workbook(temp_file.name)
    wb._external_links = []
    ws = wb.active

    # 加入 logo
    img_path = os.path.join(
        current_app.root_path,
        'material_management',
        'static',
        'logo.jpg'
    )
    img = Image(img_path)
    width_emu = int(4.09 * 360000)   # 4.09 cm
    height_emu = int(1.62 * 360000) # 1.62 cm
    marker = AnchorMarker(col=0, colOff=133350, row=1, rowOff=38100)
    size = XDRPositiveSize2D(cx=width_emu, cy=height_emu)
    img.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(img)

    # 填入 special_fields
    for field, (row, start_col, end_col, default_val) in special_fields.items():
        merged_cell = ws.cell(row=row, column=start_col)
        merged_cell.value = form_data.get(field, default_val)
        if field == "工程名稱":
            merged_cell.font = Font(size=10)

    # 填入 regular_fields
    for field, row, col in regular_fields:
        ws.cell(row=row, column=col, value=form_data.get(field, ''))

    # 填入附件欄位
    for field, row, col in attachment_fields:
        ws.cell(row=row, column=col, value=form_data.get(field, ''))

    # 處理 checkbox
    for checkboxes in [
        material_type_checkboxes,
        material_status_checkboxes,
        attachment_type_checkboxes
    ]:
        for (fld, r, c) in checkboxes:
            if form_data.get(fld):  # 若有值則勾選
                ws.cell(row=r, column=c, value="☑" + fld)
            else:
                ws.cell(row=r, column=c, value="□" + fld)

    # 填入日期
    default_date = datetime.now().strftime("%Y/%m/%d")
    ws.cell(row=21, column=7, value=form_data.get('日期', default_date))

    wb.save(temp_file.name)
    return temp_file.name
