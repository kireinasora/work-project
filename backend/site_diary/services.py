# backend/site_diary/services.py

import os
import tempfile
import shutil
import subprocess
from datetime import datetime
from typing import Dict, Any
import platform
from shutil import which

from flask import current_app
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ★ 新增：使用我們的 Mongo 物件
from backend.db import mongo


def generate_diary_xlsx_only(site_diary_doc: Dict[str, Any]) -> str:
    """
    產生/填寫 daily_report.xlsx 後，回傳填好的暫存 XLSX 檔路徑 (含所有工作表)。
    目前只用 Sheet1/Sheet2，但整份 workbook 都會被保留。
    """
    # 先依 project_id 找出對應的專案資料 (Mongo)
    project_id = site_diary_doc.get("project_id")
    project_doc = mongo.db.projects.find_one({"id": project_id})
    if not project_doc:
        raise ValueError("Project not found in MongoDB (id={}).".format(project_id))

    # 日報日期
    report_date = site_diary_doc.get("report_date")
    # Mongodb 中該欄可能是 datetime 或 None
    date_str = report_date.strftime("%Y-%m-%d") if report_date else ""

    # 專案基本資訊
    project_name = project_doc.get("name", "")
    project_job_number = project_doc.get("job_number", "")
    contractor_name = project_doc.get("contractor", "")

    # 日數 / 摘要 / 天氣
    day_count_str = str(site_diary_doc.get("day_count", ""))
    summary_str = site_diary_doc.get("summary", "")
    weather_morning = site_diary_doc.get("weather_morning", "")
    weather_noon = site_diary_doc.get("weather_noon", "")

    # 專案開始日期
    start_date = project_doc.get("start_date")
    start_date_str = start_date.strftime("%Y-%m-%d") if start_date else ""

    # 工期
    duration_days = project_doc.get("duration_days")
    duration_type = project_doc.get("duration_type")
    if duration_days:
        if duration_type == "business":
            duration_str = f"{duration_days}工作天"
        else:
            duration_str = f"{duration_days}天"
    else:
        duration_str = ""

    # 建立 worker_map / machine_map => 對應 Excel 中的儲存格 (Sheet1)
    worker_map = {
        "地盤總管": "D19",
        "工程師":  "D20",
        "管工":    "D21",
        "平水員":  "D22",
        "燒焊焊工": "D23",
        "機手":    "D24",
        "泥水工":  "D25",
        "紮鐵工":  "D26",
        "木板工":  "D27", 
        "電工":    "D28",
        "水喉工":  "D29",
        "雜工":    "D30",
    }
    machine_map = {
        "挖掘機": "G19",
        "發電機": "G20",
        "風機":   "G21",
        "泥頭車": "G22",
        "吊機":   "G23",
        "機炮":   "G24",
        "屈鐵機": "G25",  
        "風車鋸": "G26",  
    }

    # 從 site_diary_doc 取出 workers/machines 這兩個 dict
    workers_dict = site_diary_doc.get("workers", {})
    machines_dict = site_diary_doc.get("machines", {})

    # 找到 daily_report.xlsx 模板位置
    template_path = os.path.join(
        current_app.root_path,
        'site_diary',
        'templates',
        'daily_report.xlsx'
    )
    if not os.path.isfile(template_path):
        raise FileNotFoundError(f"daily_report.xlsx not found: {template_path}")

    # 在系統臨時資料夾中建立暫存路徑
    temp_dir = tempfile.mkdtemp(prefix="diary_xlsx_")

    # 用報表日期組出輸出檔名
    date_str_for_filename = "noDate" if not date_str else report_date.strftime("%Y%m%d")
    filled_xlsx_path = os.path.join(
        temp_dir,
        f"{date_str_for_filename}_daily_report.xlsx"
    )

    # 複製模板
    shutil.copy(template_path, filled_xlsx_path)

    # 打開並填寫
    wb = load_workbook(filled_xlsx_path)
    # 目前預期兩個 sheet: 「每日施工進度報告表」「每日本地工人及外地勞工施工人員紀錄表」

    # Sheet1: "每日施工進度報告表"
    if "每日施工進度報告表" in wb.sheetnames:
        sh1 = wb["每日施工進度報告表"]

        # 基本欄位
        sh1["C2"].value = date_str
        sh1["C2"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["K2"].value = start_date_str
        sh1["K2"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["D4"].value = project_name
        sh1["D4"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["D5"].value = project_job_number
        sh1["D5"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["D6"].value = contractor_name
        sh1["D6"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["E7"].value = weather_morning
        sh1["E7"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["H7"].value = weather_noon
        sh1["H7"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["D8"].value = duration_str
        sh1["D8"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["J8"].value = day_count_str
        sh1["J8"].alignment = Alignment(horizontal='center', vertical='center')

        sh1["B10"].value = summary_str
        sh1["B10"].alignment = Alignment(horizontal='center', vertical='center')

        # 工人數量
        for w_type, cell_ref in worker_map.items():
            qty = workers_dict.get(w_type, 0)
            sh1[cell_ref].value = str(qty) if qty else ""
            sh1[cell_ref].alignment = Alignment(horizontal='center', vertical='center')

        # 機器數量
        for m_type, cell_ref in machine_map.items():
            qty = machines_dict.get(m_type, 0)
            sh1[cell_ref].value = str(qty) if qty else ""
            sh1[cell_ref].alignment = Alignment(horizontal='center', vertical='center')

        # 列印範圍
        sh1.print_area = "A1:L46"

    # Sheet2: "每日本地工人及外地勞工施工人員紀錄表"
    if "每日本地工人及外地勞工施工人員紀錄表" in wb.sheetnames:
        sh2 = wb["每日本地工人及外地勞工施工人員紀錄表"]

        # B3 放日期 (用 YYYY/MM/DD)
        b3_date = report_date.strftime("%Y/%m/%d") if report_date else ""
        sh2["B3"].value = b3_date
        sh2["B3"].alignment = Alignment(horizontal='center', vertical='center')

        # 專案名稱
        sh2["B6"].value = project_name
        sh2["B6"].alignment = Alignment(horizontal='center', vertical='center')

        # 工程編號
        sh2["B7"].value = project_job_number
        sh2["B7"].alignment = Alignment(horizontal='center', vertical='center')

        # 承建商
        sh2["B8"].value = contractor_name
        sh2["B8"].alignment = Alignment(horizontal='center', vertical='center')

        # ===== 關鍵需求：若全部工人 + 機器皆為 0，則 B11-D25, B28-C32 全填 "☐"，並 C33=0, C34=0 =====
        total_workers_sum = sum(workers_dict.values())
        total_machines_sum = sum(machines_dict.values())
        if (total_workers_sum + total_machines_sum) == 0:
            # B11 ~ D25
            for row in range(11, 26):
                for col in range(2, 5):  # 2~4 => B, C, D
                    sh2.cell(row=row, column=col, value="☐")
            # B28 ~ C32
            for row in range(28, 33):
                for col in range(2, 4):  # 2~3 => B, C
                    sh2.cell(row=row, column=col, value="☐")
            # C33 = 0, C34 = 0
            sh2["C33"].value = 0
            sh2["C34"].value = 0

        sh2.print_area = "A1:E40"

    wb.save(filled_xlsx_path)
    wb.close()

    return filled_xlsx_path


def generate_diary_pdf_sheet(site_diary_doc: Dict[str, Any], sheet_name: str) -> str:
    """
    依照 sheet_name ('sheet1' or 'sheet2')，只輸出對應工作表的 PDF:
      - sheet1 => "每日施工進度報告表"
      - sheet2 => "每日本地工人及外地勞工施工人員紀錄表"

    內部流程：
      1) 先 generate_diary_xlsx_only() 產生含所有工作表的 XLSX
      2) 複製一份，刪掉不必要的工作表，只保留目標表
      3) 用 LibreOffice CLI 轉成 PDF (指定內嵌字型、PDF/A-1) 以避免中文亂碼
      4) 回傳轉好的 PDF 路徑
    """
    # 產生整本 XLSX
    xlsx_path = generate_diary_xlsx_only(site_diary_doc)
    temp_dir = os.path.dirname(xlsx_path)

    # 報表日期（命名用）
    report_date = site_diary_doc.get("report_date")
    date_str_for_filename = "noDate"
    if report_date:
        date_str_for_filename = report_date.strftime("%Y%m%d")

    # 決定要保留的工作表名稱
    if sheet_name == 'sheet1':
        target_sheet_name = "每日施工進度報告表"
        pdf_filename = f"{date_str_for_filename}每日施工進度報告表.pdf"
    else:
        target_sheet_name = "每日本地工人及外地勞工施工人員紀錄表"
        pdf_filename = f"{date_str_for_filename}施工人員紀錄表.pdf"

    pdf_path = os.path.join(temp_dir, pdf_filename)

    # 先另存一份只保留指定工作表的 XLSX
    single_sheet_xlsx = os.path.join(temp_dir, f"{date_str_for_filename}_{sheet_name}_only.xlsx")
    shutil.copy(xlsx_path, single_sheet_xlsx)

    wb = load_workbook(single_sheet_xlsx)
    if target_sheet_name not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"指定工作表 '{target_sheet_name}' 不存在於模板中。")

    # 刪除其他工作表
    for sn in wb.sheetnames:
        if sn != target_sheet_name:
            wb.remove(wb[sn])
    wb.save(single_sheet_xlsx)
    wb.close()

    # 執行 LibreOffice CLI 轉 PDF
    lo_exec = get_libreoffice_cmd()
    input_filter = "Calc Office Open XML"

    # ★ 為了避免中文亂碼，指定 PDF/A-1 與 EmbedStandardFonts=true
    pdf_filter_options = "SelectPdfVersion=1;EmbedStandardFonts=true"
    output_filter = f"pdf:calc_pdf_Export:{pdf_filter_options}"

    if platform.system().lower().startswith("win"):
        # Windows系統
        cmd = (
            f'"{lo_exec}" --headless --convert-to "{output_filter}" '
            f'--infilter="{input_filter}" "{single_sheet_xlsx}" --outdir "{temp_dir}"'
        )
        try:
            result = subprocess.run(cmd, shell=True, check=True, capture_output=True)
        except subprocess.CalledProcessError as e:
            err_msg = e.stderr.decode('utf-8', errors='replace') if e.stderr else str(e)
            raise RuntimeError(f"LibreOffice PDF conversion failed (Windows): {err_msg}") from e
    else:
        # Linux / Mac
        cmd_list = [
            lo_exec, "--headless",
            "--convert-to", output_filter,
            f'--infilter={input_filter}',
            single_sheet_xlsx,
            "--outdir", temp_dir
        ]
        try:
            result = subprocess.run(cmd_list, shell=False, check=True, capture_output=True)
        except subprocess.CalledProcessError as e:
            err_msg = e.stderr.decode('utf-8', errors='replace') if e.stderr else str(e)
            raise RuntimeError(f"LibreOffice PDF conversion failed: {err_msg}") from e

    generated_pdf_name = os.path.splitext(single_sheet_xlsx)[0] + ".pdf"
    if os.path.isfile(generated_pdf_name):
        os.rename(generated_pdf_name, pdf_path)
    else:
        raise RuntimeError(f"PDF 轉檔失敗: {generated_pdf_name} 不存在")

    if not os.path.isfile(pdf_path):
        raise RuntimeError(f"PDF 轉檔失敗: {pdf_path} 不存在")

    return pdf_path


def get_libreoffice_cmd() -> str:
    """
    取得 LibreOffice 執行檔路徑:
      1) 若有環境變數 LIBREOFFICE_PATH，先檢查是否有效
      2) 否則嘗試 which('soffice') 或 which('libreoffice')
      3) 全部找不到就拋錯
    """
    custom_path = os.environ.get("LIBREOFFICE_PATH")
    if custom_path:
        if os.path.isfile(custom_path):
            return custom_path
        soffice_path = os.path.join(custom_path, "soffice")
        libreoffice_path = os.path.join(custom_path, "libreoffice")
        if os.path.isfile(soffice_path):
            return soffice_path
        if os.path.isfile(libreoffice_path):
            return libreoffice_path
        raise RuntimeError(f"LIBREOFFICE_PATH 無效: {custom_path}")

    for candidate in ("soffice", "libreoffice"):
        exe_path = which(candidate)
        if exe_path:
            return exe_path

    raise RuntimeError(
        "無法自動找到 libreoffice/soffice！請安裝 LibreOffice 或設定 LIBREOFFICE_PATH"
    )
