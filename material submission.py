import tkinter as tk
from tkinter import ttk, filedialog
from tkcalendar import DateEntry
from openpyxl import load_workbook
from openpyxl.styles import Font
import shutil
import os
from datetime import datetime
import openpyxl

class MaterialSubmissionApp:
    def __init__(self, master):
        self.master = master
        master.title("Material Submission Form")
        master.geometry("1000x800")
        master.configure(bg="#f0f0f0")

        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TFrame", background="#f0f0f0")
        self.style.configure("TLabel", background="#f0f0f0", font=("Arial", 10))
        self.style.configure("TEntry", font=("Arial", 10))
        self.style.configure("TCheckbutton", background="#f0f0f0", font=("Arial", 10))
        self.style.configure("TButton", font=("Arial", 10, "bold"))

        self.template_file = "材料報批表.xlsx"
        self.entries = {}
        self.checkboxes = {}
        self.special_fields = {
            "工程編號": (6, 2, 4, "37/2024/DVPS"),
            "工程名稱": (7, 2, 4, "黑沙馬路行人道優化工程(第二期)"),
            "文件編號": (6, 8, 8, "") 
        }
        self.regular_fields = [
            ("報批之材料", 11, 3),
            ("牌子(如有)", 12, 3),
            ("預算表之項目編號", 11, 7),
            ("型號", 12, 6),
            ("貨期", 13, 6),
            ("數量", 14, 6),
            ("附件", 15, 6)
        ]
        self.material_type_checkboxes = [
            ("結構", 7, 6),
            ("供水", 8, 6),
            ("建築", 7, 8),
            ("電氣", 8, 8),
            ("排水", 7, 10),
            ("其他", 8, 10)
        ]
        self.material_status_checkboxes = [
            ("與設計相同", 13, 1),
            ("與標書相同", 14, 1),
            ("與後加工程建議書相同", 15, 1),
            ("同等質量", 16, 1),
            ("替換材料", 17, 1),
            ("原設計沒有指定", 18, 1)
        ]
        self.attachment_type_checkboxes = [
            ("樣板", 16, 5),
            ("目錄", 17, 5),
            ("來源證", 16, 7),
            ("其他", 17, 7)
        ]

        self.create_widgets()

    def create_widgets(self):
        main_frame = ttk.Frame(self.master, padding="20 20 20 20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(main_frame, bg="#f0f0f0")
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Create input fields for special and regular fields
        for field, (_, _, _, default_value) in self.special_fields.items():
            self.create_entry_field(field, default_value)

        for field, _, _ in self.regular_fields:
            self.create_entry_field(field)

        # Create checkbox groups
        self.create_checkbox_group("Material Type", self.material_type_checkboxes, columns=3)
        self.create_checkbox_group("Material Status", self.material_status_checkboxes, columns=1)
        self.create_checkbox_group("Attachment Type", self.attachment_type_checkboxes, columns=2)

        # Create date field
        self.create_entry_field("日期", is_date=True)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Save button
        button_frame = ttk.Frame(self.scrollable_frame)
        button_frame.pack(fill=tk.X, pady=10)
        self.save_button = ttk.Button(button_frame, text="Save As", command=self.save_form, style="TButton")
        self.save_button.pack()

    def create_entry_field(self, field, default_value=None, is_date=False):
        frame = ttk.Frame(self.scrollable_frame)
        label = ttk.Label(frame, text=f"{field}:", width=20, anchor="e")
        label.pack(side=tk.LEFT, padx=(0, 10))

        if is_date:
            entry = DateEntry(frame, width=20, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy/mm/dd')
            entry.set_date(datetime.now())
            self.date_entry = entry
        else:
            entry = ttk.Entry(frame, width=60)
            if default_value:
                entry.insert(0, default_value)
            self.entries[field] = entry

        entry.pack(side=tk.LEFT, expand=True, fill=tk.X)
        frame.pack(fill=tk.X, padx=10, pady=5)

    def create_checkbox_group(self, title, checkboxes, columns):
        group_frame = ttk.Frame(self.scrollable_frame)
        group_frame.pack(fill=tk.X, padx=10, pady=5)

        label = ttk.Label(group_frame, text=f"{title}:", width=20, anchor="e")
        label.pack(side=tk.LEFT, padx=(0, 10))

        checkbox_frame = ttk.Frame(group_frame)
        checkbox_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        for i, (field, _, _) in enumerate(checkboxes):
            var = tk.BooleanVar(value=False)
            checkbox = ttk.Checkbutton(checkbox_frame, text=field, variable=var)
            checkbox.grid(row=i//columns, column=i%columns, sticky="w", padx=5, pady=2)
            self.checkboxes[field] = var

    def save_form(self):
        try:
            # Ask user where to save the file
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )

            if not save_path:
                return  # User cancelled the save operation

            # Copy the template file to the new location
            shutil.copy2(self.template_file, save_path)

            # Load the copied file
            wb = load_workbook(save_path)
            wb._external_links = []
            ws = wb.active

            # Fill in the special fields
            for field, (row, start_col, end_col, _) in self.special_fields.items():
                merged_cell = ws.cell(row=row, column=start_col)
                merged_cell.value = self.entries[field].get()
                if field == "工程名稱":
                    merged_cell.font = Font(size=10)
                
            # Fill in the regular fields
            for field, row, col in self.regular_fields:
                ws.cell(row=row, column=col, value=self.entries[field].get())

            # Fill in the material type checkboxes
            for field, row, col in self.material_type_checkboxes:
                ws.cell(row=row, column=col, value="☑" + field if self.checkboxes[field].get() else "□" + field)

            # Fill in the material status checkboxes
            for field, row, col in self.material_status_checkboxes:
                ws.cell(row=row, column=col, value="☑" + field if self.checkboxes[field].get() else "□" + field)

            # Fill in the attachment type checkboxes
            for field, row, col in self.attachment_type_checkboxes:
                ws.cell(row=row, column=col, value="☑" + field if self.checkboxes[field].get() else "□" + field)

            # Fill in the date
            ws.cell(row=21, column=7, value=self.date_entry.get_date().strftime("%Y/%m/%d"))

            # Save the workbook
            wb.save(save_path)

            print("Form saved successfully as Excel!")

        except Exception as e:
            print(f"An error occurred while saving: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = MaterialSubmissionApp(root)
    root.mainloop()
